"""
app.py
------
VocaShot 백엔드 (FastAPI).

기능 흐름:
  회원가입/이메일 인증/로그인  →  단어장 PDF 업로드 & OCR 추출(진행바)
  →  단어장 저장(이름/설명)  →  단어 범위 선택 & 시험지 생성(진행바)
  →  시험지 저장  →  시험 응시(시간 제한 옵션)  →  AI 채점 & 결과 저장(이력)

모든 데이터는 SQLite(database.py)에 사용자별로 저장된다.
"""

import os
import json
import shutil
import threading
import secrets
import string
from datetime import datetime, timezone

from fastapi import FastAPI, File, UploadFile, Form, Request, Response, HTTPException, Depends
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import database as db
import auth
import progress
import ai_extract
import ai_quiz
import pdf_export
import srs
import planner

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="VocaShot")


@app.middleware("http")
async def _no_cache_static(request: Request, call_next):
    """정적 파일(HTML/JS/CSS)은 항상 재검증하도록 no-cache 헤더 부여.
    (배포 후 브라우저가 옛 app.js/index.html 을 캐시해 쓰는 문제 방지)"""
    resp = await call_next(request)
    path = request.url.path
    if path == "/" or path.endswith((".html", ".js", ".css")):
        resp.headers["Cache-Control"] = "no-cache, must-revalidate"
    return resp

# ------------------------------------------------------------------
# 회원 등급별 한도
#   basic(기본)  : 무료 사용자 기본 한도
#   premium(프리미엄): 사실상 무제한에 가까운 넉넉한 한도
# 결제 시스템은 추후 연동 예정 — 지금은 /api/upgrade 로 전환(개발/데모용).
# ------------------------------------------------------------------
# AI 크레딧: 비싼 AI 작업(시험지 생성/예문 첨삭/플래너 추천)을 크레딧으로 계량해
# 원가를 통제한다. Gemini Flash 1회 호출 원가는 수원(₩) 미만이라, 아래 크레딧 한도
# 내에서는 구독료 대비 확실히 흑자다(경쟁사 Vocaro 대비 기능 우위로 가격 방어).
CREDIT_COST = {"quiz": 4, "sentence": 1, "planner": 3, "study": 1, "grade": 1}
# 크레딧을 소모하는(프리미엄) 학습·게임 모드. 기본 모드는 무료(0).
PREMIUM_STUDY = {"cover", "dictation", "speed", "spelling", "match_def", "sentence_fill"}

LIMITS = {
    "basic": {
        "wordbooks": 3, "exams": 10, "extract_pages": 10, "quiz_questions": 15,
        "can_create_class": False, "usage_check": False, "max_classes": 0, "max_students": 0,
        "ai_credits": 50, "pdf_custom": False, "pdf_watermark": True,
        "label": "무료", "price": "₩0",
    },
    "premium": {
        "wordbooks": 9999, "exams": 9999, "extract_pages": 50, "quiz_questions": 50,
        "can_create_class": False, "usage_check": True, "max_classes": 0, "max_students": 0,
        "ai_credits": 600, "pdf_custom": True, "pdf_watermark": False,
        "label": "프리미엄", "price": "₩4,900/월",
    },
    "teacher": {
        "wordbooks": 9999, "exams": 9999, "extract_pages": 100, "quiz_questions": 100,
        "can_create_class": True, "usage_check": True, "max_classes": 5, "max_students": 40,
        "ai_credits": 2500, "pdf_custom": True, "pdf_watermark": False,
        "label": "선생님 Basic", "price": "₩12,900/월",
    },
    "teacher_pro": {
        "wordbooks": 9999, "exams": 9999, "extract_pages": 200, "quiz_questions": 100,
        "can_create_class": True, "usage_check": True, "max_classes": 20, "max_students": 60,
        "ai_credits": 6000, "pdf_custom": True, "pdf_watermark": False, "priority_support": True,
        "label": "선생님 Pro", "price": "₩29,900/월",
    },
}

# 애드온(일회성 구매) 카탈로그: kind → 단위·증가량·가격(데모)
ADDONS = {
    "credits":  {"label": "AI 크레딧 팩", "unit": "500 크레딧", "amount": 500, "price": "₩4,900"},
    "classes":  {"label": "반 슬롯", "unit": "반 +1", "amount": 1, "price": "₩3,900"},
    "students": {"label": "학생 슬롯", "unit": "학생 +10", "amount": 10, "price": "₩3,900"},
}


def limits_for(tier: str):
    return LIMITS.get(tier or "basic", LIMITS["basic"])


def effective_limits(user_id: int):
    """등급 기본 한도 + 애드온(추가 크레딧·반·학생 슬롯)을 합친 실질 한도."""
    row = db.query_one(
        "SELECT tier, COALESCE(addon_credits,0) ac, COALESCE(addon_classes,0) acl, "
        "COALESCE(addon_students,0) ast FROM users WHERE id=?", (user_id,))
    base = dict(limits_for((row["tier"] if row else "basic")))
    if row:
        base["ai_credits"] = base.get("ai_credits", 0) + row["ac"]
        base["max_classes"] = base.get("max_classes", 0) + (row["acl"] if base.get("max_classes", 0) else 0)
        base["max_students"] = base.get("max_students", 0) + (row["ast"] if base.get("max_students", 0) else 0)
        base["addon_credits"] = row["ac"]
        base["addon_classes"] = row["acl"]
        base["addon_students"] = row["ast"]
    return base


@app.on_event("startup")
def _startup():
    db.init_db()
    auth.ensure_test_account()
    print("✅ VocaShot 서버 준비 완료. http://127.0.0.1:8000")


def _now():
    return datetime.now(timezone.utc).isoformat()


def _notify(user_id: int, title: str, body: str = "", ntype: str = "info"):
    """사용자에게 알림 한 건 추가."""
    db.execute(
        "INSERT INTO notifications (user_id, type, title, body, is_read, created_at) "
        "VALUES (?, ?, ?, ?, 0, ?)",
        (user_id, ntype, title, body, _now()), commit=True)


def _ai_usage(user_id: int, feature: str = "credits"):
    """이번 달 누적 사용량(크레딧)."""
    ym = datetime.now(timezone.utc).strftime("%Y-%m")
    row = db.query_one("SELECT count FROM ai_usage WHERE user_id=? AND ym=? AND feature=?",
                       (user_id, ym, feature))
    return row["count"] if row else 0


def _credit_status(user):
    """AI 크레딧 현황(애드온 포함). 반환: (used, limit, remaining)."""
    lim = effective_limits(user["id"]).get("ai_credits", 50)
    used = _ai_usage(user["id"], "credits")
    return used, lim, max(0, lim - used)


def _charge_credits(user, cost: int):
    """크레딧을 차감(사용). 잔액 부족이면 False, 충분하면 차감 후 True."""
    used, lim, remaining = _credit_status(user)
    if remaining < cost:
        return False
    ym = datetime.now(timezone.utc).strftime("%Y-%m")
    db.execute(
        "INSERT INTO ai_usage (user_id, ym, feature, count) VALUES (?,?,?,?) "
        "ON CONFLICT(user_id, ym, feature) DO UPDATE SET count=count+?",
        (user["id"], ym, "credits", cost, cost), commit=True)
    return True


def _charge_credits_by_id(uid: int, cost: int):
    """특정 사용자(uid)의 크레딧 차감. 잔액 부족이면 False."""
    lim = effective_limits(uid).get("ai_credits", 50)
    used = _ai_usage(uid, "credits")
    if lim - used < cost:
        return False
    ym = datetime.now(timezone.utc).strftime("%Y-%m")
    db.execute(
        "INSERT INTO ai_usage (user_id, ym, feature, count) VALUES (?,?,?,?) "
        "ON CONFLICT(user_id, ym, feature) DO UPDATE SET count=count+?",
        (uid, ym, "credits", cost, cost), commit=True)
    return True


# ==================================================================
# 인증 의존성
# ==================================================================
def current_user(request: Request):
    """세션 쿠키로 로그인 사용자 확인. 없으면 401."""
    token = request.cookies.get("session")
    user = auth.get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    return user


# ==================================================================
# 인증 API
# ==================================================================
class SignupBody(BaseModel):
    email: str
    password: str
    tier: str = "basic"   # basic 또는 teacher(선생님 회원, 유료)


class VerifyBody(BaseModel):
    email: str
    code: str


class LoginBody(BaseModel):
    email: str
    password: str


@app.post("/api/signup")
def signup(body: SignupBody):
    email = body.email.strip().lower()
    if "@" not in email:
        raise HTTPException(400, "올바른 이메일 형식이 아닙니다.")
    pw_err = auth.validate_password(body.password)
    if pw_err:
        raise HTTPException(400, pw_err)

    existing = auth.get_user_by_email(email)
    if existing and existing["is_verified"]:
        raise HTTPException(400, "이미 가입된 이메일입니다.")

    # 선생님 회원은 유료(데모에서는 결제 없이 생성). 그 외는 기본 회원.
    signup_tier = "teacher" if body.tier == "teacher" else "basic"

    # 미인증 상태로 계정 생성(또는 비밀번호 갱신) 후 인증 코드 발송
    if not existing:
        auth.create_user(email, body.password, verified=False, tier=signup_tier)
    else:
        auth.set_tier(existing["id"], signup_tier)  # 재가입 시 선택 등급 반영
    code = auth.issue_verification_code(email)
    auth.send_verification_email(email, code)

    resp = {"status": "success", "message": "인증 코드를 이메일로 보냈습니다."}
    if auth.DEV_MODE:
        # 개발 모드: 테스트 편의를 위해 코드 노출 (운영에서는 DEV_MODE=0)
        resp["dev_code"] = code
    return resp


@app.post("/api/verify")
def verify(body: VerifyBody, response: Response):
    email = body.email.strip().lower()
    if not auth.check_verification_code(email, body.code):
        raise HTTPException(400, "인증 코드가 올바르지 않거나 만료되었습니다.")
    auth.mark_verified(email)
    user = auth.get_user_by_email(email)
    token = auth.create_session(user["id"])
    response.set_cookie("session", token, httponly=True, samesite="lax", max_age=60 * 60 * 24 * 30)
    return {"status": "success", "email": email}


@app.post("/api/resend-code")
def resend_code(body: SignupBody):
    email = body.email.strip().lower()
    if not auth.get_user_by_email(email):
        raise HTTPException(400, "먼저 회원가입을 해주세요.")
    code = auth.issue_verification_code(email)
    auth.send_verification_email(email, code)
    resp = {"status": "success"}
    if auth.DEV_MODE:
        resp["dev_code"] = code
    return resp


@app.post("/api/login")
def login(body: LoginBody, response: Response):
    email = body.email.strip().lower()
    user = auth.get_user_by_email(email)
    if not user or not auth.verify_password(body.password, user["password_hash"]):
        raise HTTPException(400, "이메일 또는 비밀번호가 올바르지 않습니다.")
    if not user["is_verified"]:
        raise HTTPException(403, "이메일 인증이 필요합니다.")
    token = auth.create_session(user["id"])
    response.set_cookie("session", token, httponly=True, samesite="lax", max_age=60 * 60 * 24 * 30)
    return {"status": "success", "email": email}


@app.post("/api/logout")
def logout(request: Request, response: Response):
    auth.delete_session(request.cookies.get("session"))
    response.delete_cookie("session")
    return {"status": "success"}


@app.get("/api/me")
def me(user=Depends(current_user)):
    full = auth.get_user_full(user["id"])
    tier = full.get("tier") or "basic"
    lim = effective_limits(user["id"])
    # 현재 사용량 (한도 표시용)
    wb_used = db.query_one("SELECT COUNT(*) c FROM wordbooks WHERE user_id=?", (user["id"],))["c"]
    ex_used = db.query_one("SELECT COUNT(*) c FROM exams WHERE user_id=?", (user["id"],))["c"]
    return {
        "id": full["id"],
        "email": full["email"],
        "nickname": full.get("nickname") or full["email"].split("@")[0],
        "avatar": full.get("avatar") or "",
        "tier": tier,
        "limits": lim,
        "explain_lang": full.get("explain_lang") or "ko",
        "usage": {"wordbooks": wb_used, "exams": ex_used,
                  "ai_credits": _ai_usage(user["id"], "credits")},
        "credit_cost": CREDIT_COST,
    }


class StudyChargeBody(BaseModel):
    mode: str


@app.post("/api/study/charge")
def study_charge(body: StudyChargeBody, user=Depends(current_user)):
    """프리미엄 학습·게임 모드 시작 시 크레딧 차감. 기본 모드는 0."""
    cost = CREDIT_COST["study"] if body.mode in PREMIUM_STUDY else 0
    if cost and not _charge_credits(user, cost):
        used, limit, remaining = _credit_status(user)
        raise HTTPException(
            402, f"AI 크레딧이 부족해요. (남은 크레딧 {remaining}) 프리미엄으로 업그레이드하면 이 학습 모드를 마음껏 쓸 수 있어요.")
    used, limit, remaining = _credit_status(user)
    return {"status": "success", "cost": cost, "remaining": remaining, "premium": body.mode in PREMIUM_STUDY}


class PrefsBody(BaseModel):
    explain_lang: str


@app.post("/api/profile/prefs")
def update_prefs(body: PrefsBody, user=Depends(current_user)):
    """문제·해설 기본 언어 설정."""
    lang = body.explain_lang if body.explain_lang in ("ko", "en", "ja", "zh", "es", "fr", "de") else "ko"
    auth.set_explain_lang(user["id"], lang)
    return {"status": "success", "explain_lang": lang}


class ProfileBody(BaseModel):
    nickname: str


@app.post("/api/profile")
def update_profile(body: ProfileBody, user=Depends(current_user)):
    nickname = body.nickname.strip()
    if not (1 <= len(nickname) <= 20):
        raise HTTPException(400, "닉네임은 1~20자여야 합니다.")
    auth.update_profile(user["id"], nickname=nickname)
    return {"status": "success", "nickname": nickname}


@app.post("/api/profile/avatar")
def upload_avatar(user=Depends(current_user), file: UploadFile = File(...)):
    """프로필 사진을 data URL(base64)로 변환해 DB에 저장 (작은 이미지 가정)."""
    ext = os.path.splitext(file.filename)[1].lower()
    mime = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".webp": "image/webp"}.get(ext)
    if not mime:
        raise HTTPException(400, "이미지 파일(PNG/JPG/GIF/WEBP)만 지원합니다.")
    raw = file.file.read()
    if len(raw) > 2 * 1024 * 1024:
        raise HTTPException(400, "이미지는 2MB 이하만 업로드할 수 있습니다.")
    import base64
    data_url = f"data:{mime};base64,{base64.b64encode(raw).decode()}"
    auth.update_profile(user["id"], avatar=data_url)
    return {"status": "success", "avatar": data_url}


class AvatarDataBody(BaseModel):
    avatar: str  # data:image/...;base64,... (클라이언트에서 크롭/리사이즈된 이미지)


@app.post("/api/profile/avatar-crop")
def save_cropped_avatar(body: AvatarDataBody, user=Depends(current_user)):
    """클라이언트에서 크롭·리사이즈한 이미지(data URL)를 그대로 저장."""
    data = body.avatar
    if not data.startswith("data:image/"):
        raise HTTPException(400, "올바른 이미지 데이터가 아닙니다.")
    if len(data) > 700_000:  # 256px 정사각형이면 넉넉한 상한
        raise HTTPException(400, "이미지 용량이 너무 큽니다.")
    auth.update_profile(user["id"], avatar=data)
    return {"status": "success", "avatar": data}


class PasswordBody(BaseModel):
    current: str
    new: str


@app.post("/api/profile/password")
def change_password(body: PasswordBody, user=Depends(current_user)):
    """비밀번호 변경: 현재 비밀번호 확인 후 새 비밀번호로 교체."""
    full = db.query_one("SELECT password_hash FROM users WHERE id=?", (user["id"],))
    if not full or not auth.verify_password(body.current, full["password_hash"]):
        raise HTTPException(400, "현재 비밀번호가 올바르지 않습니다.")
    pw_err = auth.validate_password(body.new)
    if pw_err:
        raise HTTPException(400, pw_err)
    if body.current == body.new:
        raise HTTPException(400, "새 비밀번호가 기존과 동일합니다.")
    auth.update_password(user["id"], body.new)
    return {"status": "success"}


@app.get("/api/stats")
def get_stats(user=Depends(current_user)):
    """대시보드용 학습 통계 집계."""
    uid = user["id"]
    wb = db.query_one("SELECT COUNT(*) c FROM wordbooks WHERE user_id=?", (uid,))["c"]
    words = db.query_one(
        "SELECT COUNT(*) c FROM words w JOIN wordbooks wb ON wb.id=w.wordbook_id WHERE wb.user_id=?",
        (uid,))["c"]
    ex = db.query_one("SELECT COUNT(*) c FROM exams WHERE user_id=?", (uid,))["c"]
    agg = db.query_one(
        "SELECT COUNT(*) c, COALESCE(ROUND(AVG(score)),0) avg, COALESCE(MAX(score),0) best "
        "FROM attempts WHERE user_id=?", (uid,))
    # 시험지별 성적 (최고/평균/응시횟수) — 대시보드 위젯용
    exam_scores = db.query_all(
        "SELECT e.id AS exam_id, e.name, COUNT(a.id) AS attempts, "
        "  MAX(a.score) AS best, ROUND(AVG(a.score)) AS avg "
        "FROM exams e JOIN attempts a ON a.exam_id=e.id AND a.user_id=? "
        "WHERE e.user_id=? GROUP BY e.id ORDER BY MAX(a.created_at) DESC LIMIT 8", (uid, uid))

    # 이번 주 학습 요약 (최근 7일 응시 수)
    week_ago = (datetime.now(timezone.utc) - _timedelta(days=7)).isoformat()
    week_attempts = db.query_one(
        "SELECT COUNT(*) c FROM attempts WHERE user_id=? AND created_at >= ?", (uid, week_ago))["c"]
    # 최근 응시 기록 (최신순 6개)
    recent = db.query_all(
        "SELECT a.id, a.score, a.correct, a.total, a.created_at, e.name AS exam_name "
        "FROM attempts a JOIN exams e ON e.id=a.exam_id WHERE a.user_id=? "
        "ORDER BY a.created_at DESC LIMIT 6", (uid,))
    return {
        "totals": {"wordbooks": wb, "words": words, "exams": ex, "attempts": agg["c"]},
        "avg_score": agg["avg"],
        "best_score": agg["best"],
        "exam_scores": exam_scores,
        "week_attempts": week_attempts,
        "recent": recent,
        "due_total": srs.due_count(uid),
        "mastery": srs.analysis(uid)["distribution"],
    }


@app.get("/api/calendar")
def get_calendar(user=Depends(current_user)):
    """
    캘린더/망각곡선 뷰용 데이터.
    - study: 날짜별 응시 기록(횟수/평균점수)  → 학습 기록
    - reviews: 날짜별 복습 예정 단어 수(next_review) → 망각곡선 기반 복습 알림
    """
    uid = user["id"]
    # 학습 기록: 응시일(로컬 아닌 UTC 날짜 기준, ISO 앞 10자)
    study_rows = db.query_all(
        "SELECT substr(created_at,1,10) d, COUNT(*) attempts, ROUND(AVG(score)) avg "
        "FROM attempts WHERE user_id=? GROUP BY d", (uid,))
    study = {r["d"]: {"attempts": r["attempts"], "avg": r["avg"]} for r in study_rows}

    # 복습 예정: next_review 날짜별 단어 수
    review_rows = db.query_all(
        "SELECT substr(next_review,1,10) d, COUNT(*) cnt FROM word_mastery "
        "WHERE user_id=? AND next_review IS NOT NULL GROUP BY d", (uid,))
    reviews = {r["d"]: r["cnt"] for r in review_rows if r["d"]}

    # 사용자 커스텀 일정(메모): 날짜별 항목 목록 (드래그로 날짜 이동 가능)
    note_rows = db.query_all(
        "SELECT id, date, text, done FROM calendar_notes WHERE user_id=? ORDER BY id", (uid,))
    note_items = {}
    for r in note_rows:
        note_items.setdefault(r["date"], []).append({"id": r["id"], "text": r["text"], "done": r["done"]})
    notes = {d: len(v) for d, v in note_items.items()}

    return {"study": study, "reviews": reviews, "notes": notes,
            "note_items": note_items, "today": _now()[:10]}


@app.get("/api/calendar/day/{date}")
def get_calendar_day(date: str, user=Depends(current_user)):
    """특정 날짜의 상세: 응시 기록 + 복습 예정 단어 + 사용자 커스텀 일정."""
    uid = user["id"]
    attempts = db.query_all(
        "SELECT a.id, a.score, a.correct, a.total, a.created_at, e.name AS exam_name "
        "FROM attempts a JOIN exams e ON e.id=a.exam_id "
        "WHERE a.user_id=? AND substr(a.created_at,1,10)=? ORDER BY a.created_at DESC", (uid, date))
    due = db.query_all(
        "SELECT word, meaning FROM word_mastery WHERE user_id=? AND substr(next_review,1,10)=? LIMIT 60",
        (uid, date))
    notes = db.query_all(
        "SELECT id, text, done FROM calendar_notes WHERE user_id=? AND date=? ORDER BY id", (uid, date))
    return {"date": date, "attempts": attempts, "due": due, "notes": notes}


class CalNoteBody(BaseModel):
    date: str
    text: str


@app.post("/api/calendar/note")
def add_calendar_note(body: CalNoteBody, user=Depends(current_user)):
    if not body.text.strip():
        raise HTTPException(400, "내용을 입력해주세요.")
    nid = db.execute(
        "INSERT INTO calendar_notes (user_id, date, text, done, created_at) VALUES (?,?,?,0,?)",
        (user["id"], body.date[:10], body.text.strip(), _now()), commit=True)
    return {"status": "success", "id": nid}


@app.post("/api/calendar/note/{note_id}/toggle")
def toggle_calendar_note(note_id: int, user=Depends(current_user)):
    n = db.query_one("SELECT done FROM calendar_notes WHERE id=? AND user_id=?", (note_id, user["id"]))
    if not n:
        raise HTTPException(404, "일정을 찾을 수 없습니다.")
    db.execute("UPDATE calendar_notes SET done=? WHERE id=?", (0 if n["done"] else 1, note_id), commit=True)
    return {"status": "success"}


class CalMoveBody(BaseModel):
    date: str


@app.post("/api/calendar/note/{note_id}/move")
def move_calendar_note(note_id: int, body: CalMoveBody, user=Depends(current_user)):
    """커스텀 일정을 다른 날짜로 이동(드래그 앤 드롭)."""
    n = db.query_one("SELECT id FROM calendar_notes WHERE id=? AND user_id=?", (note_id, user["id"]))
    if not n:
        raise HTTPException(404, "일정을 찾을 수 없습니다.")
    db.execute("UPDATE calendar_notes SET date=? WHERE id=?", (body.date[:10], note_id), commit=True)
    return {"status": "success"}


@app.delete("/api/calendar/note/{note_id}")
def delete_calendar_note(note_id: int, user=Depends(current_user)):
    db.execute("DELETE FROM calendar_notes WHERE id=? AND user_id=?", (note_id, user["id"]), commit=True)
    return {"status": "success"}


class SrsRescheduleBody(BaseModel):
    from_date: str
    to_date: str


@app.post("/api/srs/reschedule")
def srs_reschedule(body: SrsRescheduleBody, user=Depends(current_user)):
    """망각곡선 복습 칩을 드래그로 다른 날짜에 재배치(해당 날짜 예정 단어들의 next_review 이동)."""
    frm, to = body.from_date[:10], body.to_date[:10]
    if frm == to:
        return {"status": "success", "moved": 0}
    n = db.query_one(
        "SELECT COUNT(*) c FROM word_mastery WHERE user_id=? AND substr(next_review,1,10)=?",
        (user["id"], frm))["c"]
    if not n:
        raise HTTPException(400, "옮길 복습이 없습니다.")
    db.execute("UPDATE word_mastery SET next_review=? WHERE user_id=? AND substr(next_review,1,10)=?",
               (to, user["id"], frm), commit=True)
    _notify(user["id"], "망각곡선 복습을 옮겼어요 🔁",
            f"{frm}의 복습 {n}개를 {to}로 옮겼어요. 그날 복습하면 망각곡선이 다시 계산됩니다.", ntype="info")
    return {"status": "success", "moved": n}


# ==================================================================
# 학습 플래너 (망각곡선 기반)
# ==================================================================
from datetime import date as _date, timedelta as _timedelta


def _words_for_wordbooks(user_id, wb_ids):
    """선택한 단어장들(본인 소유)의 단어를 모두 모아 반환."""
    words = []
    for wid in wb_ids:
        wb = db.query_one("SELECT id FROM wordbooks WHERE id=? AND user_id=?", (wid, user_id))
        if not wb:
            continue
        rows = db.query_all("SELECT word, meaning, reading, example, definition FROM words WHERE wordbook_id=? ORDER BY seq", (wid,))
        for r in rows:
            r["wordbook_id"] = wid
            words.append(r)
    return words


class PlannerPreviewBody(BaseModel):
    wordbook_ids: list = []
    goal: str = ""
    period_days: int | None = None          # 목표 기간(일). 없으면 무기한
    exclude_weekdays: list = []             # 공부 안 하는 요일(JS getDay: 일0~토6)
    daily_new: int | None = None


def _parse_date(s):
    try:
        return _date.fromisoformat(s[:10]) if s else None
    except Exception:
        return None


def _exclude_py(js_days):
    """JS 요일(일0~토6) → 파이썬 weekday(월0~일6)로 변환한 집합."""
    out = set()
    for j in (js_days or []):
        try:
            out.add((int(j) - 1) % 7)
        except Exception:
            pass
    return out


@app.post("/api/planner/preview")
def planner_preview(body: PlannerPreviewBody, user=Depends(current_user)):
    """플랜 미리보기: AI 추천 파라미터 + 망각곡선 스케줄 생성(저장 X)."""
    words = _words_for_wordbooks(user["id"], body.wordbook_ids)
    if not words:
        raise HTTPException(400, "단어장을 하나 이상 선택해주세요.")
    start = _date.today()
    period = body.period_days if (body.period_days and body.period_days > 0) else None
    target = start + _timedelta(days=period) if period else None
    days_until = period
    exclude_py = _exclude_py(body.exclude_weekdays)

    # AI 맞춤 추천은 크레딧 소모(3). 잔액 부족 시 규칙 기반으로 자동 대체.
    cost = CREDIT_COST["planner"]
    charged = _charge_credits(user, cost)
    rec = planner.ai_recommend(body.goal, len(words), days_until, use_ai=charged)
    used, limit, remaining = _credit_status(user)
    allowed = charged
    daily_new = body.daily_new or rec["daily_new"]
    built = planner.build_schedule(words, start, daily_new, target_date=target,
                                   reviews=rec["reviews"], exclude_weekdays=exclude_py)
    stats = planner.plan_stats(built["schedule"])
    summary_days = [{"date": d["date"], "new": len(d["new"]), "review": len(d["review"])}
                    for d in built["schedule"]]
    return {
        "word_count": len(words),
        "daily_new": built["daily_new"],
        "intervals": built["intervals"],
        "period_days": period,
        "exclude_weekdays": body.exclude_weekdays or [],
        "target_date": target.isoformat() if target else None,
        "summary": rec["summary"],
        "ai_used": rec["source"] == "ai",
        "ai_locked": not allowed,               # 크레딧 부족 → AI 잠금
        "credits_used": used, "credits_limit": limit, "credits_left": remaining,
        "stats": stats,
        "days": summary_days,
        "retention_curve": [{"day": d, "retention": round(planner.retention_after(d) * 100)}
                            for d in [0, 1, 3, 7, 17, 39]],
    }


class PlannerSaveBody(BaseModel):
    name: str = "나의 학습 플랜"
    goal: str = ""
    wordbook_ids: list = []
    period_days: int | None = None
    exclude_weekdays: list = []
    daily_new: int | None = None
    reviews: int = 5


@app.post("/api/planner")
def planner_save(body: PlannerSaveBody, user=Depends(current_user)):
    words = _words_for_wordbooks(user["id"], body.wordbook_ids)
    if not words:
        raise HTTPException(400, "단어장을 하나 이상 선택해주세요.")
    start = _date.today()
    period = body.period_days if (body.period_days and body.period_days > 0) else None
    target = start + _timedelta(days=period) if period else None
    days_until = period
    exclude_py = _exclude_py(body.exclude_weekdays)
    daily_new = body.daily_new or planner.ai_recommend(body.goal, len(words), days_until)["daily_new"]
    built = planner.build_schedule(words, start, daily_new, target_date=target,
                                   reviews=body.reviews, exclude_weekdays=exclude_py)
    # 기존 활성 플랜 비활성화 (활성 플랜 1개 유지)
    db.execute("UPDATE study_plans SET active=0 WHERE user_id=?", (user["id"],), commit=True)
    pid = db.execute(
        "INSERT INTO study_plans (user_id, name, goal, start_date, target_date, daily_new, "
        "wordbook_ids, schedule, intervals, exclude_weekdays, active, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,1,?)",
        (user["id"], body.name.strip() or "나의 학습 플랜", body.goal,
         start.isoformat(), target.isoformat() if target else None, built["daily_new"],
         json.dumps(body.wordbook_ids), json.dumps(built["schedule"], ensure_ascii=False),
         json.dumps(built["intervals"]), json.dumps(body.exclude_weekdays or []), _now()), commit=True)
    return {"status": "success", "plan_id": pid}


@app.get("/api/planner")
def planner_get(user=Depends(current_user)):
    """활성 플랜 + 진행 상황."""
    plan = db.query_one("SELECT * FROM study_plans WHERE user_id=? AND active=1 ORDER BY created_at DESC LIMIT 1", (user["id"],))
    if not plan:
        return {"plan": None}
    schedule = json.loads(plan["schedule"])
    done = {(r["date"], r["kind"]) for r in db.query_all(
        "SELECT date, kind FROM plan_done WHERE plan_id=?", (plan["id"],))}
    total = len(schedule)
    done_days = len({d for (d, k) in done})
    # 기간(일) 역산 — 수정 화면 프리필용
    period_days = None
    if plan["target_date"] and plan["start_date"]:
        try:
            period_days = (_date.fromisoformat(plan["target_date"][:10])
                           - _date.fromisoformat(plan["start_date"][:10])).days
        except Exception:
            period_days = None
    try:
        exclude_weekdays = json.loads(plan["exclude_weekdays"] or "[]")
    except Exception:
        exclude_weekdays = []
    return {"plan": {
        "id": plan["id"], "name": plan["name"], "goal": plan["goal"],
        "start_date": plan["start_date"], "target_date": plan["target_date"],
        "period_days": period_days, "exclude_weekdays": exclude_weekdays,
        "daily_new": plan["daily_new"], "intervals": json.loads(plan["intervals"] or "[]"),
        "wordbook_ids": json.loads(plan["wordbook_ids"] or "[]"),
        "schedule": schedule, "done": list(done), "total_days": total, "done_days": done_days,
    }}


@app.delete("/api/planner/{plan_id}")
def planner_delete(plan_id: int, user=Depends(current_user)):
    p = db.query_one("SELECT id FROM study_plans WHERE id=? AND user_id=?", (plan_id, user["id"]))
    if not p:
        raise HTTPException(404, "플랜을 찾을 수 없습니다.")
    db.execute("DELETE FROM study_plans WHERE id=?", (plan_id,), commit=True)
    return {"status": "success"}


class PlanDoneBody(BaseModel):
    date: str
    kind: str  # new / review


@app.post("/api/planner/{plan_id}/complete")
def planner_complete(plan_id: int, body: PlanDoneBody, user=Depends(current_user)):
    p = db.query_one("SELECT id FROM study_plans WHERE id=? AND user_id=?", (plan_id, user["id"]))
    if not p:
        raise HTTPException(404, "플랜을 찾을 수 없습니다.")
    db.execute("INSERT OR IGNORE INTO plan_done (plan_id, user_id, date, kind) VALUES (?,?,?,?)",
               (plan_id, user["id"], body.date[:10], body.kind), commit=True)
    return {"status": "success"}


class PlanMoveBody(BaseModel):
    from_date: str
    to_date: str
    kind: str  # new / review


@app.post("/api/planner/{plan_id}/move")
def planner_move(plan_id: int, body: PlanMoveBody, user=Depends(current_user)):
    """AI 플랜의 특정 날짜(new/review) 항목을 다른 날짜로 드래그 이동."""
    p = db.query_one("SELECT * FROM study_plans WHERE id=? AND user_id=?", (plan_id, user["id"]))
    if not p:
        raise HTTPException(404, "플랜을 찾을 수 없습니다.")
    kind = body.kind if body.kind in ("new", "review") else "new"
    frm, to = body.from_date[:10], body.to_date[:10]
    if frm == to:
        return {"status": "success"}
    schedule = json.loads(p["schedule"])
    src = next((d for d in schedule if d["date"] == frm), None)
    if not src or not src.get(kind):
        raise HTTPException(400, "옮길 항목이 없습니다.")
    moving = src[kind]
    src[kind] = []
    dst = next((d for d in schedule if d["date"] == to), None)
    if not dst:
        dst = {"date": to, "new": [], "review": []}
        schedule.append(dst)
    seen = {w.get("word") for w in dst.get(kind, [])}
    dst[kind] = dst.get(kind, []) + [w for w in moving if w.get("word") not in seen]

    intervals = json.loads(p["intervals"] or "[]")
    try:
        exclude_py = _exclude_py(json.loads(p["exclude_weekdays"] or "[]"))
    except Exception:
        exclude_py = set()

    def _next_ok(d):
        while d.weekday() in exclude_py:
            d += _timedelta(days=1)
        return d

    def _merge_review(date_str, revs):
        day = next((d for d in schedule if d["date"] == date_str), None)
        if not day:
            day = {"date": date_str, "new": [], "review": []}
            schedule.append(day)
        seen_r = {w.get("word") for w in day.get("review", [])}
        day["review"] = day.get("review", []) + [w for w in revs if w.get("word") not in seen_r]

    moved = {w.get("word") for w in moving}
    recalculated = False
    if kind == "new":
        # 새 단어 학습일을 옮기면 → 그 단어들의 망각곡선 복습을 새 intro(to) 기준으로 전체 재계산
        for d in schedule:
            if d.get("review"):
                d["review"] = [w for w in d["review"] if w.get("word") not in moved]
        to_obj = _date.fromisoformat(to)
        for w in moving:
            for iv in intervals:
                rday = _next_ok(to_obj + _timedelta(days=iv)).isoformat()
                _merge_review(rday, [{"word": w.get("word", ""), "meaning": w.get("meaning", ""),
                                      "wordbook_id": w.get("wordbook_id")}])
        recalculated = True
    elif kind == "review":
        # 복습을 옮기면 → 그 뒤에 이어지는 같은 단어들의 복습도 같은 폭(delta)만큼 함께 이동(망각곡선 유지)
        delta = (_date.fromisoformat(to) - _date.fromisoformat(frm)).days
        shifted = {}
        for d in schedule:
            if d["date"] > frm and d.get("review"):
                move_ws = [w for w in d["review"] if w.get("word") in moved]
                if move_ws:
                    d["review"] = [w for w in d["review"] if w.get("word") not in moved]
                    nd = _next_ok(_date.fromisoformat(d["date"]) + _timedelta(days=delta)).isoformat()
                    shifted.setdefault(nd, []).extend(move_ws)
        for date_str, revs in shifted.items():
            _merge_review(date_str, revs)
        recalculated = bool(shifted)

    # 신규/복습이 모두 빈 날짜는 스케줄에서 제거
    schedule = [d for d in schedule if (d.get("new") or d.get("review"))]
    schedule.sort(key=lambda d: d["date"])
    db.execute("UPDATE study_plans SET schedule=? WHERE id=?",
               (json.dumps(schedule, ensure_ascii=False), plan_id), commit=True)
    # 완료 기록도 함께 이동
    db.execute("UPDATE plan_done SET date=? WHERE plan_id=? AND date=? AND kind=?",
               (to, plan_id, frm, kind), commit=True)
    # 조정 알림
    if recalculated:
        _notify(user["id"], "학습 플랜이 조정됐어요 📅",
                f"새 단어 학습을 {frm} → {to}로 옮기고, 망각곡선 복습 일정을 다시 계산했어요.", ntype="info")
    else:
        _notify(user["id"], "복습 일정을 옮겼어요 🔁",
                f"{frm}의 복습을 {to}로 이동했습니다.", ntype="info")
    return {"status": "success", "recalculated": recalculated}


@app.get("/api/today")
def today_tasks(user=Depends(current_user)):
    """오늘 할 일 모음: 망각곡선 복습 + 플랜의 오늘 항목 + 미응시 배정 시험."""
    uid = user["id"]
    today = _date.today().isoformat()

    # 1) 망각곡선 복습 대상 (지난 시험/복습에서 쌓인 단어)
    due = srs.due_count(uid)

    # 2) 활성 플랜의 오늘 항목
    plan = db.query_one("SELECT * FROM study_plans WHERE user_id=? AND active=1 ORDER BY created_at DESC LIMIT 1", (uid,))
    plan_today = None
    if plan:
        schedule = json.loads(plan["schedule"])
        day = next((d for d in schedule if d["date"] == today), None)
        done = {(r["date"], r["kind"]) for r in db.query_all(
            "SELECT date, kind FROM plan_done WHERE plan_id=?", (plan["id"],))}
        if day:
            plan_today = {
                "plan_id": plan["id"], "plan_name": plan["name"],
                "new": day["new"], "review": day["review"],
                "new_done": (today, "new") in done,
                "review_done": (today, "review") in done,
            }

    # 3) 내가 속한 반의 배정 시험 중 미응시
    assignments = db.query_all(
        "SELECT DISTINCT e.id AS exam_id, e.name AS exam_name, c.name AS class_name "
        "FROM class_members m JOIN assignments a ON a.class_id=m.class_id "
        "JOIN exams e ON e.id=a.exam_id JOIN classes c ON c.id=a.class_id "
        "WHERE m.student_id=? AND NOT EXISTS "
        "  (SELECT 1 FROM attempts at WHERE at.exam_id=e.id AND at.user_id=?)", (uid, uid))

    # 4) 오늘 날짜의 사용자 커스텀 일정(미완료)
    notes = db.query_all(
        "SELECT id, text FROM calendar_notes WHERE user_id=? AND date=? AND done=0 ORDER BY id", (uid, today))

    return {"date": today, "reviews_due": due, "plan_today": plan_today,
            "assignments": assignments, "notes": notes}


@app.post("/api/stats/reset")
def reset_stats(user=Depends(current_user)):
    """학습 통계 초기화: 응시 기록 + 단어 숙련도를 모두 삭제(단어장·시험지는 유지)."""
    uid = user["id"]
    db.execute("DELETE FROM attempts WHERE user_id=?", (uid,), commit=True)
    db.execute("DELETE FROM word_mastery WHERE user_id=?", (uid,), commit=True)
    return {"status": "success"}


class TierBody(BaseModel):
    tier: str


@app.post("/api/tier")
def set_tier(body: TierBody, user=Depends(current_user)):
    """등급 전환 (basic/premium/teacher). 데모용 — 실제 결제 연동은 추후."""
    if body.tier not in LIMITS:
        raise HTTPException(400, "알 수 없는 등급입니다.")
    # 로스터로 등록된 학생 계정은 등급 전환 불가(선생님 위장·무료 확산 방지)
    row = db.query_one("SELECT provider FROM users WHERE id=?", (user["id"],))
    if row and row["provider"] == "roster":
        raise HTTPException(403, "학생 계정은 등급을 변경할 수 없습니다.")
    auth.set_tier(user["id"], body.tier)
    return {"status": "success", "tier": body.tier}


class AddonBody(BaseModel):
    kind: str      # credits / classes / students
    qty: int = 1


@app.get("/api/addons")
def list_addons():
    return {"addons": ADDONS}


@app.post("/api/addons/buy")
def buy_addon(body: AddonBody, user=Depends(current_user)):
    """애드온 구매(데모: 실제 결제 없이 한도 즉시 증가)."""
    a = ADDONS.get(body.kind)
    if not a:
        raise HTTPException(400, "알 수 없는 애드온입니다.")
    row = db.query_one("SELECT provider FROM users WHERE id=?", (user["id"],))
    if row and row["provider"] == "roster":
        raise HTTPException(403, "학생 계정은 구매할 수 없습니다.")
    qty = max(1, min(int(body.qty or 1), 50))
    col = {"credits": "addon_credits", "classes": "addon_classes", "students": "addon_students"}[body.kind]
    db.execute(f"UPDATE users SET {col}=COALESCE({col},0)+? WHERE id=?",
               (a["amount"] * qty, user["id"]), commit=True)
    return {"status": "success", "kind": body.kind, "added": a["amount"] * qty,
            "limits": effective_limits(user["id"])}


# ==================================================================
# 알림
# ==================================================================
@app.get("/api/notifications")
def list_notifications(user=Depends(current_user)):
    rows = db.query_all(
        "SELECT id, type, title, body, is_read, created_at FROM notifications "
        "WHERE user_id=? ORDER BY created_at DESC LIMIT 30", (user["id"],))
    unread = db.query_one(
        "SELECT COUNT(*) c FROM notifications WHERE user_id=? AND is_read=0", (user["id"],))["c"]
    return {"notifications": rows, "unread": unread}


@app.post("/api/notifications/read")
def mark_notifications_read(user=Depends(current_user)):
    db.execute("UPDATE notifications SET is_read=1 WHERE user_id=?", (user["id"],), commit=True)
    return {"status": "success"}


@app.delete("/api/notifications")
def clear_notifications(user=Depends(current_user)):
    db.execute("DELETE FROM notifications WHERE user_id=?", (user["id"],), commit=True)
    return {"status": "success"}


# ==================================================================
# 진행 상황 폴링
# ==================================================================
@app.get("/api/progress/{job_id}")
def get_progress(job_id: str, user=Depends(current_user)):
    job = progress.get(job_id)
    if not job:
        raise HTTPException(404, "작업을 찾을 수 없습니다.")
    return job


# ==================================================================
# 단어장: 업로드 & 추출
# ==================================================================
@app.post("/api/upload")
def upload_file(user=Depends(current_user), file: UploadFile = File(...)):
    """파일을 서버 uploads/ 에 저장하고 페이지 수를 반환."""
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".pdf", ".jpg", ".jpeg", ".png"):
        raise HTTPException(400, "PDF 또는 이미지 파일만 지원합니다.")
    safe_name = f"u{user['id']}_{int(datetime.now().timestamp())}{ext}"
    dest = os.path.join(UPLOAD_DIR, safe_name)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    pages = ai_extract.count_pdf_pages(dest) if ext == ".pdf" else 1
    return {"status": "success", "filename": safe_name,
            "original_name": file.filename, "page_count": pages}


class ExtractBody(BaseModel):
    filename: str
    start_page: int = 1
    end_page: int = 1
    language: str = "en"
    meaning_lang: str = "ko"


@app.post("/api/extract-start")
def extract_start(body: ExtractBody, user=Depends(current_user)):
    """추출 작업을 백그라운드로 시작하고 job_id 반환 (진행바용)."""
    path = os.path.join(UPLOAD_DIR, os.path.basename(body.filename))
    if not os.path.exists(path):
        raise HTTPException(404, "업로드된 파일을 찾을 수 없습니다.")

    # 등급별 1회 추출 페이지 한도 검사
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    requested = max(1, body.end_page - body.start_page + 1)
    if requested > lim["extract_pages"]:
        raise HTTPException(
            403, f"{full.get('tier','basic')} 등급은 한 번에 최대 {lim['extract_pages']}페이지까지 "
                 f"추출할 수 있습니다. (요청: {requested}페이지)")

    job_id = progress.create_job()

    def report(percent, message, log):
        progress.update(job_id, percent=percent, message=message, log=log)

    def worker():
        try:
            words = ai_extract.extract_words(path, body.start_page, body.end_page, report,
                                             body.language, body.meaning_lang)
            if not words:
                progress.fail(job_id, "추출된 단어가 없습니다.")
                return
            progress.finish(job_id, {"words": words})
        except Exception as e:
            progress.fail(job_id, str(e))

    threading.Thread(target=worker, daemon=True).start()
    return {"status": "success", "job_id": job_id}


class SaveWordbookBody(BaseModel):
    name: str
    description: str = ""
    source_name: str = ""
    words: list
    language: str = "en"
    meaning_lang: str = "ko"


@app.post("/api/wordbooks")
def save_wordbook(body: SaveWordbookBody, user=Depends(current_user)):
    """추출된 단어를 단어장으로 저장."""
    if not body.name.strip():
        raise HTTPException(400, "단어장 이름을 입력해주세요.")
    if not body.words:
        raise HTTPException(400, "저장할 단어가 없습니다.")

    # 등급별 단어장 개수 한도 검사
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    used = db.query_one("SELECT COUNT(*) c FROM wordbooks WHERE user_id=?", (user["id"],))["c"]
    if used >= lim["wordbooks"]:
        raise HTTPException(
            403, f"{full.get('tier','basic')} 등급의 단어장 한도({lim['wordbooks']}개)에 도달했습니다. "
                 f"프리미엄으로 업그레이드하면 무제한으로 저장할 수 있습니다.")

    lang = body.language if body.language in ("en", "zh", "ja", "es", "fr", "de") else "en"
    mlang = body.meaning_lang if body.meaning_lang in ("ko", "en", "ja", "zh", "es", "fr", "de") else "ko"
    wb_id = db.execute(
        "INSERT INTO wordbooks (user_id, name, description, source_name, created_at, language, meaning_lang) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (user["id"], body.name.strip(), body.description.strip(), body.source_name, _now(), lang, mlang),
        commit=True,
    )
    rows = [
        (wb_id, i + 1, w.get("word", ""), w.get("reading", ""), w.get("meaning", ""),
         w.get("definition", ""), w.get("example", ""), int(w.get("page", 1) or 1))
        for i, w in enumerate(body.words)
    ]
    db.executemany(
        "INSERT INTO words (wordbook_id, seq, word, reading, meaning, definition, example, page) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    return {"status": "success", "wordbook_id": wb_id, "word_count": len(rows)}


@app.get("/api/wordbooks")
def list_wordbooks(user=Depends(current_user)):
    return {"wordbooks": db.query_all(
        "SELECT wb.id, wb.name, wb.description, wb.source_name, wb.created_at, wb.language, "
        "  (SELECT COUNT(*) FROM words w WHERE w.wordbook_id = wb.id) AS word_count, "
        "  (SELECT COUNT(*) FROM exams e WHERE e.wordbook_id = wb.id) AS exam_count "
        "FROM wordbooks wb WHERE wb.user_id = ? ORDER BY wb.created_at DESC",
        (user["id"],),
    )}


@app.get("/api/wordbooks/{wb_id}")
def get_wordbook(wb_id: int, user=Depends(current_user)):
    wb = db.query_one("SELECT * FROM wordbooks WHERE id=? AND user_id=?", (wb_id, user["id"]))
    if not wb:
        raise HTTPException(404, "단어장을 찾을 수 없습니다.")
    words = db.query_all("SELECT * FROM words WHERE wordbook_id=? ORDER BY seq", (wb_id,))
    src = wb.get("source_name") or ""
    wb["has_source"] = bool(src) and os.path.exists(os.path.join(UPLOAD_DIR, os.path.basename(src)))
    return {"wordbook": wb, "words": words}


@app.get("/api/wordbooks/{wb_id}/source")
def download_wordbook_source(wb_id: int, user=Depends(current_user)):
    """단어장의 원본 업로드 파일(PDF/이미지) 다운로드."""
    wb = db.query_one("SELECT name, source_name FROM wordbooks WHERE id=? AND user_id=?", (wb_id, user["id"]))
    if not wb:
        raise HTTPException(404, "단어장을 찾을 수 없습니다.")
    src = os.path.basename(wb.get("source_name") or "")
    path = os.path.join(UPLOAD_DIR, src)
    if not src or not os.path.exists(path):
        raise HTTPException(404, "원본 파일이 없습니다.")
    ext = os.path.splitext(src)[1] or ".pdf"
    mime = {".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg"}.get(ext.lower(), "application/octet-stream")
    return FileResponse(path, filename=f"{wb['name']}_원본{ext}", media_type=mime)


@app.delete("/api/wordbooks/{wb_id}")
def delete_wordbook(wb_id: int, user=Depends(current_user)):
    wb = db.query_one("SELECT id FROM wordbooks WHERE id=? AND user_id=?", (wb_id, user["id"]))
    if not wb:
        raise HTTPException(404, "단어장을 찾을 수 없습니다.")
    # 이 단어장으로 만든 시험지도 함께 삭제 (FK 는 SET NULL 이라 명시적으로 제거)
    exams = db.query_all("SELECT id FROM exams WHERE wordbook_id=? AND user_id=?", (wb_id, user["id"]))
    for e in exams:
        db.execute("DELETE FROM exams WHERE id=?", (e["id"],), commit=True)
    db.execute("DELETE FROM wordbooks WHERE id=?", (wb_id,), commit=True)
    return {"status": "success", "deleted_exams": len(exams)}


@app.get("/api/wordbooks/{wb_id}/exam-count")
def wordbook_exam_count(wb_id: int, user=Depends(current_user)):
    """단어장 삭제 확인용: 함께 삭제될 시험지 개수."""
    c = db.query_one("SELECT COUNT(*) c FROM exams WHERE wordbook_id=? AND user_id=?", (wb_id, user["id"]))["c"]
    return {"exam_count": c}


# ==================================================================
# 시험지: 생성 & 저장
# ==================================================================
class GenerateBody(BaseModel):
    words: list          # 선택된 단어 목록
    format: str          # toefl / sat / word_to_meaning / meaning_to_word / fill_meaning
    count: int = 10
    language: str = "en"
    meaning_lang: str = "ko"       # ko(한국어 뜻) / en(영영풀이)
    explain_lang: str | None = None  # 문제·해설 언어 (없으면 설정값)


@app.post("/api/exams/generate-start")
def generate_start(body: GenerateBody, user=Depends(current_user)):
    """문제 생성을 백그라운드로 시작하고 job_id 반환 (진행바용)."""
    if not body.words:
        raise HTTPException(400, "선택된 단어가 없습니다.")
    if body.format not in ai_quiz.FORMAT_LABELS:
        raise HTTPException(400, "알 수 없는 시험지 형식입니다.")

    # 등급별 문항 수 한도 검사
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    if body.count > lim["quiz_questions"]:
        raise HTTPException(
            403, f"{full.get('tier','basic')} 등급은 한 시험지에 최대 {lim['quiz_questions']}문항까지 "
                 f"생성할 수 있습니다. (요청: {body.count}문항)")

    # AI 크레딧 차감 (시험지 생성)
    if not _charge_credits(user, CREDIT_COST["quiz"]):
        used, climit, remaining = _credit_status(user)
        raise HTTPException(
            402, f"AI 크레딧이 부족합니다. (남은 크레딧 {remaining} · 시험지 생성 {CREDIT_COST['quiz']} 필요) "
                 f"프리미엄으로 업그레이드하면 매달 넉넉한 크레딧을 받아요.")

    explain_lang = body.explain_lang or full.get("explain_lang") or "ko"
    job_id = progress.create_job()

    def report(percent, message, log):
        progress.update(job_id, percent=percent, message=message, log=log)

    def worker():
        try:
            quiz = ai_quiz.generate_quiz(body.words, body.format, max(1, body.count), report,
                                         body.language, body.meaning_lang, explain_lang)
            if not quiz:
                progress.fail(job_id, "문제 생성에 실패했습니다.")
                return
            progress.finish(job_id, {"questions": quiz})
        except Exception as e:
            progress.fail(job_id, str(e))

    threading.Thread(target=worker, daemon=True).start()
    return {"status": "success", "job_id": job_id}


class SaveExamBody(BaseModel):
    name: str
    format: str
    questions: list
    wordbook_id: int | None = None


@app.post("/api/exams")
def save_exam(body: SaveExamBody, user=Depends(current_user)):
    if not body.name.strip() or not body.questions:
        raise HTTPException(400, "시험지 이름과 문제가 필요합니다.")

    # 등급별 시험지 개수 한도 검사
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    used = db.query_one("SELECT COUNT(*) c FROM exams WHERE user_id=?", (user["id"],))["c"]
    if used >= lim["exams"]:
        raise HTTPException(
            403, f"{full.get('tier','basic')} 등급의 시험지 한도({lim['exams']}개)에 도달했습니다. "
                 f"프리미엄으로 업그레이드하면 무제한으로 저장할 수 있습니다.")

    exam_id = db.execute(
        "INSERT INTO exams (user_id, wordbook_id, name, format, questions, question_count, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (user["id"], body.wordbook_id, body.name.strip(), body.format,
         json.dumps(body.questions, ensure_ascii=False), len(body.questions), _now()),
        commit=True,
    )
    return {"status": "success", "exam_id": exam_id}


@app.get("/api/exams")
def list_exams(user=Depends(current_user)):
    # attempt_count / best_score 는 '현재 사용자 본인'의 응시 기준 (내 시험지 = 내 결과)
    rows = db.query_all(
        "SELECT e.id, e.name, e.format, e.question_count, e.created_at, wb.name AS wordbook_name, "
        "  (SELECT COUNT(*) FROM attempts a WHERE a.exam_id = e.id AND a.user_id = ?) AS attempt_count, "
        "  (SELECT MAX(a.score) FROM attempts a WHERE a.exam_id = e.id AND a.user_id = ?) AS best_score "
        "FROM exams e LEFT JOIN wordbooks wb ON wb.id = e.wordbook_id "
        "WHERE e.user_id = ? ORDER BY e.created_at DESC",
        (user["id"], user["id"], user["id"]),
    )
    for r in rows:
        r["format_label"] = ai_quiz.FORMAT_LABELS.get(r["format"], r["format"])
    return {"exams": rows}


def _exam_visible_to(user_id: int, exam_id: int):
    """소유한 시험지 또는 '내가 속한 반에 배정된' 시험지면 접근 허용."""
    exam = db.query_one("SELECT * FROM exams WHERE id=?", (exam_id,))
    if not exam:
        return None
    if exam["user_id"] == user_id:
        return exam
    row = db.query_one(
        "SELECT 1 FROM assignments a JOIN class_members m ON m.class_id = a.class_id "
        "WHERE a.exam_id = ? AND m.student_id = ?", (exam_id, user_id))
    return exam if row else None


@app.get("/api/exams/{exam_id}")
def get_exam(exam_id: int, user=Depends(current_user)):
    exam = _exam_visible_to(user["id"], exam_id)
    if not exam:
        raise HTTPException(404, "시험지를 찾을 수 없습니다.")
    exam["questions"] = json.loads(exam["questions"])
    exam["format_label"] = ai_quiz.FORMAT_LABELS.get(exam["format"], exam["format"])
    return {"exam": exam}


@app.delete("/api/exams/{exam_id}")
def delete_exam(exam_id: int, user=Depends(current_user)):
    exam = db.query_one("SELECT id FROM exams WHERE id=? AND user_id=?", (exam_id, user["id"]))
    if not exam:
        raise HTTPException(404, "시험지를 찾을 수 없습니다.")
    db.execute("DELETE FROM exams WHERE id=?", (exam_id,), commit=True)
    return {"status": "success"}


# ==================================================================
# 응시 & 채점
# ==================================================================
class GradeBody(BaseModel):
    exam_id: int
    answers: dict           # {"0": "답", "1": "답", ...}
    time_taken: int = 0


@app.post("/api/attempts/grade-start")
def grade_start(body: GradeBody, user=Depends(current_user)):
    """채점을 백그라운드로 시작(주관식은 AI 채점). 완료 시 응시 기록 저장."""
    exam = _exam_visible_to(user["id"], body.exam_id)  # 배정받은 학생도 응시 가능
    if not exam:
        raise HTTPException(404, "시험지를 찾을 수 없습니다.")
    quiz = json.loads(exam["questions"])

    job_id = progress.create_job()

    def report(percent, message, log):
        progress.update(job_id, percent=percent, message=message, log=log)

    # 로스터 학생(공유 링크 응시)의 채점은 담당 선생님 크레딧에서 차감
    roster = db.query_one("SELECT teacher_id FROM roster_students WHERE user_id=?", (user["id"],))

    def worker():
        try:
            report(10, "채점 시작...", "📝 답안 채점을 시작합니다.")
            outcome = ai_quiz.grade(quiz, body.answers, report)
            attempt_id = db.execute(
                "INSERT INTO attempts (exam_id, user_id, score, correct, total, time_taken, results, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (body.exam_id, user["id"], outcome["score"], outcome["correct"],
                 outcome["total"], body.time_taken,
                 json.dumps(outcome["results"], ensure_ascii=False), _now()),
                commit=True,
            )
            outcome["attempt_id"] = attempt_id
            # 단어별 숙련도 갱신 (SRS 복습 스케줄 + 취약 분석)
            srs.record_results(user["id"], exam["wordbook_id"], quiz, outcome["results"])
            # 선생님 크레딧 차감 + 응시 알림
            if roster:
                tid = roster["teacher_id"]
                charged = _charge_credits_by_id(tid, CREDIT_COST["grade"])
                nick = (auth.get_user_full(user["id"]) or {}).get("nickname") or "학생"
                _notify(tid, "학생이 시험을 풀었어요 📝",
                        f"'{nick}' 학생이 '{exam['name']}'을 {outcome['score']}점으로 제출했어요. "
                        f"(채점 크레딧 {CREDIT_COST['grade']} 차감{'' if charged else ' 실패 — 크레딧 부족'})",
                        ntype="assign")
            progress.finish(job_id, outcome)
        except Exception as e:
            progress.fail(job_id, str(e))

    threading.Thread(target=worker, daemon=True).start()
    return {"status": "success", "job_id": job_id}


# ==================================================================
# 취약 단어 분석 & SRS 복습
# ==================================================================
@app.get("/api/analysis")
def get_analysis(wordbook_id: int = None, user=Depends(current_user)):
    base = srs.analysis(user["id"], wordbook_id)
    # 학습중 / 취약 단어를 분리해서 각각 목록 제공 (대시보드 미니 리스트용)
    params = [user["id"]]
    wb_filter = ""
    if wordbook_id:
        wb_filter = "AND wordbook_id=? "
        params.append(wordbook_id)
    rows = db.query_all(
        f"SELECT word, meaning, wordbook_id, correct_count, wrong_count, streak, last_result "
        f"FROM word_mastery WHERE user_id=? {wb_filter}", tuple(params))
    learning, weak = [], []
    for r in rows:
        lvl = srs.mastery_level(r["streak"], r["last_result"])
        total = r["correct_count"] + r["wrong_count"]
        acc = round(r["correct_count"] / total * 100) if total else 0
        item = {"word": r["word"], "meaning": r["meaning"], "wordbook_id": r["wordbook_id"],
                "wrong": r["wrong_count"], "correct": r["correct_count"],
                "accuracy": acc, "level": lvl, "streak": r["streak"]}
        if lvl == "weak":
            weak.append(item)
        elif lvl == "learning":
            learning.append(item)
    weak.sort(key=lambda x: (-x["wrong"], x["accuracy"]))
    learning.sort(key=lambda x: (-x["streak"], -x["correct"]))
    base["weak_words"] = weak[:40]
    base["learning_words"] = learning[:40]
    return base


@app.get("/api/review/due")
def review_due(wordbook_id: int = None, count: int = 15, user=Depends(current_user)):
    """
    복습할 단어로 로컬 4지선다(단어→뜻) 덱을 구성 (AI 호출 없음).
    오답 보기는 같은 사용자의 단어들에서 가져온다.
    """
    due = srs.due_words(user["id"], wordbook_id, limit=count)
    if not due:
        return {"questions": [], "due_total": srs.due_count(user["id"])}

    # 오답 보기 풀: 사용자의 모든 단어 뜻
    pool = db.query_all(
        "SELECT DISTINCT w.word, w.meaning FROM words w "
        "JOIN wordbooks wb ON wb.id=w.wordbook_id WHERE wb.user_id=? AND w.meaning != ''",
        (user["id"],),
    )
    import random
    questions = []
    for d in due:
        answer = d["meaning"] or "(뜻 없음)"
        distractors = [p["meaning"] for p in pool
                       if p["meaning"] and p["meaning"] != answer]
        random.shuffle(distractors)
        opts = [answer] + distractors[:3]
        random.shuffle(opts)
        questions.append({
            "type": "mc",
            "question": d["word"],
            "options": opts,
            "answer": answer,
            "explanation": "",
            "word": d["word"],
            "meaning": answer,
            "wordbook_id": d["wordbook_id"],
        })
    return {"questions": questions, "due_total": srs.due_count(user["id"])}


class ReviewGradeBody(BaseModel):
    items: list  # [{wordbook_id, word, meaning, correct}]


class SentenceCheckBody(BaseModel):
    word: str
    meaning: str = ""
    sentence: str
    language: str = "en"


@app.post("/api/learn/check-sentence")
def check_sentence(body: SentenceCheckBody, user=Depends(current_user)):
    """플래시카드 예문 AI 첨삭. AI 크레딧 1 소모(무료 등급도 크레딧 내에서 사용 가능)."""
    if not body.sentence.strip():
        raise HTTPException(400, "예문을 입력해주세요.")
    if not _charge_credits(user, CREDIT_COST["sentence"]):
        used, climit, remaining = _credit_status(user)
        raise HTTPException(
            402, f"AI 크레딧이 부족합니다. (남은 크레딧 {remaining}) 프리미엄으로 업그레이드하면 넉넉한 크레딧을 받아요.")
    ex_lang = auth.get_user_full(user["id"]).get("explain_lang") or "ko"
    return ai_quiz.check_usage(body.word, body.meaning, body.sentence, body.language, ex_lang)


@app.post("/api/review/grade")
def review_grade(body: ReviewGradeBody, user=Depends(current_user)):
    """복습 결과를 숙련도에 반영."""
    correct = 0
    for it in body.items:
        ok = bool(it.get("correct"))
        correct += 1 if ok else 0
        srs.update_mastery(user["id"], int(it.get("wordbook_id") or 0),
                           it.get("word", ""), it.get("meaning", ""), ok)
    total = len(body.items)
    return {"status": "success", "correct": correct, "total": total,
            "score": round(correct / total * 100) if total else 0,
            "due_total": srs.due_count(user["id"])}


# ==================================================================
# 학부모 리포트 (읽기전용 공유 링크)
# ==================================================================
def _build_report(uid: int):
    """공유용 진도 리포트 데이터 집계."""
    u = auth.get_user_full(uid)
    if not u:
        return None
    wb = db.query_one("SELECT COUNT(*) c FROM wordbooks WHERE user_id=?", (uid,))["c"]
    words = db.query_one(
        "SELECT COUNT(*) c FROM words w JOIN wordbooks wb ON wb.id=w.wordbook_id WHERE wb.user_id=?",
        (uid,))["c"]
    ex = db.query_one("SELECT COUNT(*) c FROM exams WHERE user_id=?", (uid,))["c"]
    agg = db.query_one(
        "SELECT COUNT(*) c, COALESCE(ROUND(AVG(score)),0) avg, COALESCE(MAX(score),0) best "
        "FROM attempts WHERE user_id=?", (uid,))
    trend = list(reversed(db.query_all(
        "SELECT score, created_at FROM attempts WHERE user_id=? ORDER BY created_at DESC LIMIT 20", (uid,))))
    recent = db.query_all(
        "SELECT a.score, a.correct, a.total, a.created_at, e.name AS exam_name "
        "FROM attempts a JOIN exams e ON e.id=a.exam_id WHERE a.user_id=? "
        "ORDER BY a.created_at DESC LIMIT 8", (uid,))
    an = srs.analysis(uid)
    return {
        "nickname": u.get("nickname") or u["email"].split("@")[0],
        "generated_at": _now(),
        "totals": {"wordbooks": wb, "words": words, "exams": ex, "attempts": agg["c"]},
        "avg_score": agg["avg"], "best_score": agg["best"],
        "mastery": an["distribution"], "weak_words": an["weak_words"][:15],
        "trend": trend, "recent": recent,
    }


@app.post("/api/report/share")
def create_share(user=Depends(current_user)):
    """공유 토큰 발급(이미 있으면 재사용)."""
    row = db.query_one("SELECT token FROM report_shares WHERE user_id=?", (user["id"],))
    if row:
        token = row["token"]
    else:
        token = secrets.token_urlsafe(12)
        db.execute("INSERT INTO report_shares (token, user_id, created_at) VALUES (?, ?, ?)",
                   (token, user["id"], _now()), commit=True)
    return {"token": token, "path": f"/report.html?t={token}"}


@app.get("/api/report/share")
def get_share(user=Depends(current_user)):
    row = db.query_one("SELECT token FROM report_shares WHERE user_id=?", (user["id"],))
    return {"token": row["token"] if row else None,
            "path": f"/report.html?t={row['token']}" if row else None}


@app.delete("/api/report/share")
def revoke_share(user=Depends(current_user)):
    db.execute("DELETE FROM report_shares WHERE user_id=?", (user["id"],), commit=True)
    return {"status": "success"}


@app.get("/api/public/report/{token}")
def public_report(token: str):
    """토큰만 있으면 로그인 없이 볼 수 있는 리포트 (학부모용)."""
    row = db.query_one("SELECT user_id FROM report_shares WHERE token=?", (token,))
    if not row:
        raise HTTPException(404, "리포트를 찾을 수 없습니다.")
    data = _build_report(row["user_id"])
    if not data:
        raise HTTPException(404, "리포트를 찾을 수 없습니다.")
    return data


# ==================================================================
# 학원 · 교사 모드 (반 생성 / 참여 / 시험지 배정 / 반 대시보드)
# ==================================================================
def _gen_join_code():
    alphabet = string.ascii_uppercase + string.digits
    for _ in range(50):
        code = "".join(secrets.choice(alphabet) for _ in range(6))
        if not db.query_one("SELECT 1 FROM classes WHERE join_code=?", (code,)):
            return code
    return secrets.token_hex(3).upper()


class ClassBody(BaseModel):
    name: str


@app.post("/api/classes")
def create_class(body: ClassBody, user=Depends(current_user)):
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    if not lim.get("can_create_class"):
        raise HTTPException(403, "반 개설은 선생님 회원만 가능합니다. 프로필에서 선생님 회원으로 전환해주세요.")
    if not body.name.strip():
        raise HTTPException(400, "반 이름을 입력해주세요.")
    # 반 개수 상한 (남용 방지, 애드온 포함)
    max_classes = effective_limits(user["id"]).get("max_classes", 0)
    cur = db.query_one("SELECT COUNT(*) c FROM classes WHERE teacher_id=?", (user["id"],))["c"]
    if max_classes and cur >= max_classes:
        raise HTTPException(403, f"반은 최대 {max_classes}개까지 만들 수 있어요.")
    code = _gen_join_code()
    cid = db.execute("INSERT INTO classes (teacher_id, name, join_code, created_at) VALUES (?, ?, ?, ?)",
                     (user["id"], body.name.strip(), code, _now()), commit=True)
    return {"status": "success", "class_id": cid, "join_code": code}


# ---------- 로스터 학생 (선생님 등록 · 반 코드 + 숫자 학생 ID) ----------
def _gen_student_sid(cid):
    """반 안에서만 유일한 4자리 숫자 학생 ID."""
    for _ in range(80):
        sid = "".join(secrets.choice(string.digits) for _ in range(4))
        if not db.query_one("SELECT 1 FROM roster_students WHERE class_id=? AND sid=?", (cid, sid)):
            return sid
    # 폴백: 순번
    n = db.query_one("SELECT COUNT(*) c FROM roster_students WHERE class_id=?", (cid,))["c"]
    return str(1000 + n)


class RosterAddBody(BaseModel):
    name: str


@app.post("/api/classes/{cid}/students")
def add_roster_student(cid: int, body: RosterAddBody, user=Depends(current_user)):
    """선생님이 학생을 로스터에 등록 → 숫자 학생 ID 발급(학습용 계정 자동 생성)."""
    cls = db.query_one("SELECT id, name, join_code FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "학생 이름을 입력해주세요.")
    max_students = effective_limits(user["id"]).get("max_students", 0)
    cur = db.query_one("SELECT COUNT(*) c FROM class_members WHERE class_id=?", (cid,))["c"]
    if max_students and cur >= max_students:
        raise HTTPException(403, f"한 반의 학생은 최대 {max_students}명까지예요.")
    sid = _gen_student_sid(cid)
    # 학습/응시를 위한 백엔드 계정 (로그인은 반 코드 + 숫자 ID로)
    backing_email = f"c{cid}s{sid}@student.vocashot"
    uid = auth.create_user(backing_email, secrets.token_hex(8), verified=True, nickname=name, tier="basic")
    db.execute("UPDATE users SET provider='roster' WHERE id=?", (uid,), commit=True)
    db.execute("INSERT INTO class_members (class_id, student_id, joined_at) VALUES (?, ?, ?)",
               (cid, uid, _now()), commit=True)
    # code 컬럼은 과거 UNIQUE 제약이 있어 유일값 필요 → 반-학생ID 조합 사용(레거시 호환)
    legacy_code = f"{cid}-{sid}"
    db.execute(
        "INSERT INTO roster_students (class_id, teacher_id, user_id, code, pin, sid, name, created_at) "
        "VALUES (?,?,?,?,?,?,?,?)", (cid, user["id"], uid, legacy_code, "", sid, name, _now()), commit=True)
    return {"status": "success", "student_id": uid, "sid": sid,
            "join_code": cls["join_code"], "name": name}


@app.delete("/api/classes/{cid}/students/{sid}")
def remove_roster_student(cid: int, sid: int, user=Depends(current_user)):
    """로스터 학생 삭제(sid = 학생 user_id). 학습 계정까지 제거."""
    cls = db.query_one("SELECT id FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    r = db.query_one("SELECT id FROM roster_students WHERE user_id=? AND class_id=?", (sid, cid))
    if r:
        # 로스터 학생: 백엔드 계정까지 삭제(CASCADE로 관련 데이터 정리)
        db.execute("DELETE FROM users WHERE id=?", (sid,), commit=True)
        db.execute("DELETE FROM roster_students WHERE id=?", (r["id"],), commit=True)
    else:
        # 예전 자율 참여 학생: 반에서만 제외
        db.execute("DELETE FROM class_members WHERE class_id=? AND student_id=?", (cid, sid), commit=True)
    return {"status": "success"}


# ---------- 학생 응시(공유 링크 · 로그인 불필요) ----------
@app.get("/api/public/class/{code}")
def public_class_info(code: str):
    """공유 링크: 반 코드로 반 이름 확인(응시 게이트 표시용)."""
    cls = db.query_one("SELECT name FROM classes WHERE join_code=?", (code.strip().upper(),))
    if not cls:
        raise HTTPException(404, "링크의 반을 찾을 수 없습니다.")
    return {"class_name": cls["name"]}


class TakeEnterBody(BaseModel):
    code: str      # 공유 링크의 반 코드
    sid: str       # 선생님이 준 숫자 학생 ID
    name: str = ""  # 학생이 입력한 이름


@app.post("/api/public/class/enter")
def public_class_enter(body: TakeEnterBody, response: Response):
    """학생 응시 시작: 반 코드 + 학생 ID 확인 → 세션 발급(백엔드 계정). 로그인 화면 없이 링크로 진입."""
    code = body.code.strip().upper()
    sid = body.sid.strip()
    cls = db.query_one("SELECT id, name FROM classes WHERE join_code=?", (code,))
    if not cls:
        raise HTTPException(400, "링크의 반을 찾을 수 없습니다.")
    r = db.query_one("SELECT user_id FROM roster_students WHERE class_id=? AND sid=?", (cls["id"], sid))
    if not r:
        raise HTTPException(400, "학생 ID가 올바르지 않아요. 선생님이 준 번호를 확인하세요.")
    # 입력한 이름을 학생 표시 이름으로 반영(선생님 명단 이름 갱신)
    name = (body.name or "").strip()
    if name:
        auth.update_profile(r["user_id"], nickname=name)
    token = auth.create_session(r["user_id"])
    response.set_cookie("session", token, httponly=True, samesite="lax", max_age=60 * 60 * 24 * 30)
    return {"status": "success", "class_id": cls["id"], "class_name": cls["name"]}


@app.get("/api/classes")
def my_classes(user=Depends(current_user)):
    """내가 선생님인 반 목록."""
    return {"classes": db.query_all(
        "SELECT c.id, c.name, c.join_code, c.created_at, "
        "  (SELECT COUNT(*) FROM class_members m WHERE m.class_id=c.id) AS student_count, "
        "  (SELECT COUNT(*) FROM assignments a WHERE a.class_id=c.id) AS assignment_count "
        "FROM classes c WHERE c.teacher_id=? ORDER BY c.created_at DESC", (user["id"],))}


@app.get("/api/teacher/overview")
def teacher_overview(user=Depends(current_user)):
    """선생님 대시보드용 요약."""
    uid = user["id"]
    if not limits_for(auth.get_user_full(uid).get("tier")).get("can_create_class"):
        return {"is_teacher": False}
    classes = db.query_all(
        "SELECT c.id, c.name, c.join_code, "
        "  (SELECT COUNT(*) FROM class_members m WHERE m.class_id=c.id) AS student_count, "
        "  (SELECT COUNT(*) FROM assignments a WHERE a.class_id=c.id) AS assignment_count "
        "FROM classes c WHERE c.teacher_id=? ORDER BY c.created_at DESC", (uid,))
    # 각 반의 학생 평균(배정 시험지 최고점 평균)
    for c in classes:
        row = db.query_one(
            "SELECT ROUND(AVG(best)) avg FROM ("
            "  SELECT MAX(at.score) best FROM attempts at "
            "  JOIN assignments a ON a.exam_id=at.exam_id AND a.class_id=? "
            "  JOIN class_members m ON m.class_id=a.class_id AND m.student_id=at.user_id "
            "  GROUP BY at.user_id, at.exam_id)", (c["id"],))
        c["avg_score"] = row["avg"] if row and row["avg"] is not None else None
    total_students = db.query_one(
        "SELECT COUNT(DISTINCT m.student_id) c FROM class_members m "
        "JOIN classes cl ON cl.id=m.class_id WHERE cl.teacher_id=?", (uid,))["c"]
    total_assign = sum(c["assignment_count"] for c in classes)
    scored = [c["avg_score"] for c in classes if c["avg_score"] is not None]
    avg_all = round(sum(scored) / len(scored)) if scored else 0
    # 최근 학생 응시 활동
    recent = db.query_all(
        "SELECT u.nickname, u.email, c.name class_name, e.name exam_name, at.score, at.created_at "
        "FROM attempts at JOIN assignments a ON a.exam_id=at.exam_id "
        "JOIN classes c ON c.id=a.class_id AND c.teacher_id=? "
        "JOIN class_members m ON m.class_id=c.id AND m.student_id=at.user_id "
        "JOIN users u ON u.id=at.user_id JOIN exams e ON e.id=at.exam_id "
        "ORDER BY at.created_at DESC LIMIT 12", (uid,))
    students = _teacher_student_rows(uid, classes)
    return {"is_teacher": True, "classes": classes, "recent": recent, "students": students,
            "totals": {"classes": len(classes), "students": total_students,
                       "assignments": total_assign, "avg": avg_all}}


def _teacher_student_rows(uid, classes=None):
    """선생님의 모든 학생별 요약: 이름·반·학생ID·최근 시험(이름·점수)·평균."""
    if classes is None:
        classes = db.query_all("SELECT id, name FROM classes WHERE teacher_id=? ORDER BY created_at DESC", (uid,))
    rows = []
    for c in classes:
        members = db.query_all(
            "SELECT u.id, u.nickname, rs.sid FROM class_members m JOIN users u ON u.id=m.student_id "
            "LEFT JOIN roster_students rs ON rs.user_id=u.id AND rs.class_id=m.class_id "
            "WHERE m.class_id=? ORDER BY u.nickname", (c["id"],))
        exam_ids = [e["exam_id"] for e in db.query_all("SELECT exam_id FROM assignments WHERE class_id=?", (c["id"],))]
        for mem in members:
            recent_exam = recent_score = avg = attempts = None
            if exam_ids:
                ph = ",".join("?" * len(exam_ids))
                r = db.query_one(
                    f"SELECT e.name exam_name, at.score FROM attempts at JOIN exams e ON e.id=at.exam_id "
                    f"WHERE at.user_id=? AND at.exam_id IN ({ph}) ORDER BY at.created_at DESC LIMIT 1",
                    (mem["id"], *exam_ids))
                if r:
                    recent_exam, recent_score = r["exam_name"], r["score"]
                arow = db.query_one(
                    f"SELECT ROUND(AVG(best)) avg, COUNT(*) n FROM "
                    f"(SELECT MAX(score) best FROM attempts WHERE user_id=? AND exam_id IN ({ph}) GROUP BY exam_id)",
                    (mem["id"], *exam_ids))
                if arow and arow["avg"] is not None:
                    avg, attempts = arow["avg"], arow["n"]
            rows.append({"student_id": mem["id"], "name": mem["nickname"], "sid": mem["sid"] or "",
                         "class_id": c["id"], "class_name": c["name"],
                         "recent_exam": recent_exam, "recent_score": recent_score,
                         "avg": avg, "done_count": attempts or 0, "assigned_count": len(exam_ids)})
    return rows


@app.get("/api/teacher/students.xlsx")
def teacher_students_xlsx(user=Depends(current_user)):
    """선생님용: 학생 통계 xlsx 다운로드."""
    uid = user["id"]
    if not limits_for(auth.get_user_full(uid).get("tier")).get("can_create_class"):
        raise HTTPException(403, "선생님 회원만 사용할 수 있어요.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = _teacher_student_rows(uid)
    wb = Workbook()
    ws = wb.active
    ws.title = "학생 통계"
    headers = ["반", "학생", "학생 ID", "배정 과제", "완료", "최근 시험", "최근 점수", "평균 점수"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="2979FF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([
            r["class_name"], r["name"], r["sid"],
            r["assigned_count"], r["done_count"],
            r["recent_exam"] or "-",
            f'{r["recent_score"]}%' if r["recent_score"] is not None else "-",
            f'{r["avg"]}%' if r["avg"] is not None else "-",
        ])
    widths = [18, 14, 10, 10, 8, 24, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    out = os.path.join(UPLOAD_DIR, f"students_{uid}.xlsx")
    wb.save(out)
    fname = f"VocaShot_학생통계_{_date.today().isoformat()}.xlsx"
    return FileResponse(out, filename=fname,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# NOTE: 고정 경로(/joined)는 가변 경로(/{cid})보다 먼저 등록해야 매칭됨
@app.get("/api/classes/joined")
def joined_classes(user=Depends(current_user)):
    """내가 학생으로 참여한 반 목록."""
    return {"classes": db.query_all(
        "SELECT c.id, c.name, u.nickname AS teacher_name, "
        "  (SELECT COUNT(*) FROM assignments a WHERE a.class_id=c.id) AS assignment_count "
        "FROM class_members m JOIN classes c ON c.id=m.class_id JOIN users u ON u.id=c.teacher_id "
        "WHERE m.student_id=? ORDER BY m.joined_at DESC", (user["id"],))}


@app.get("/api/classes/joined/{cid}")
def joined_class_detail(cid: int, user=Depends(current_user)):
    """학생용: 참여한 반에 배정된 시험지 + 내 점수."""
    if not db.query_one("SELECT 1 FROM class_members WHERE class_id=? AND student_id=?", (cid, user["id"])):
        raise HTTPException(404, "참여한 반이 아닙니다.")
    cls = db.query_one(
        "SELECT c.id, c.name, u.nickname AS teacher_name FROM classes c "
        "JOIN users u ON u.id=c.teacher_id WHERE c.id=?", (cid,))
    assignments = db.query_all(
        "SELECT a.exam_id, e.name AS exam_name, e.format, e.question_count, "
        "  (SELECT MAX(score) FROM attempts at WHERE at.exam_id=a.exam_id AND at.user_id=?) AS my_best, "
        "  (SELECT COUNT(*) FROM attempts at WHERE at.exam_id=a.exam_id AND at.user_id=?) AS my_attempts "
        "FROM assignments a JOIN exams e ON e.id=a.exam_id WHERE a.class_id=? ORDER BY a.assigned_at",
        (user["id"], user["id"], cid))
    for a in assignments:
        a["format_label"] = ai_quiz.FORMAT_LABELS.get(a["format"], a["format"])
    wordbooks = db.query_all(
        "SELECT wa.wordbook_id, w.name AS wordbook_name, w.language, "
        "  (SELECT COUNT(*) FROM words wd WHERE wd.wordbook_id=w.id) AS word_count "
        "FROM wordbook_assignments wa JOIN wordbooks w ON w.id=wa.wordbook_id "
        "WHERE wa.class_id=? ORDER BY wa.assigned_at", (cid,))
    return {"class": cls, "assignments": assignments, "wordbooks": wordbooks}


@app.get("/api/classes/joined/{cid}/wordbook/{wid}/words")
def joined_wordbook_words(cid: int, wid: int, user=Depends(current_user)):
    """학생용: 반에 배정된 단어장의 단어 목록(학습용)."""
    if not db.query_one("SELECT 1 FROM class_members WHERE class_id=? AND student_id=?", (cid, user["id"])):
        raise HTTPException(404, "참여한 반이 아닙니다.")
    if not db.query_one("SELECT 1 FROM wordbook_assignments WHERE class_id=? AND wordbook_id=?", (cid, wid)):
        raise HTTPException(404, "이 반에 배정된 단어장이 아닙니다.")
    words = db.query_all(
        "SELECT word, meaning, reading, example, definition FROM words WHERE wordbook_id=? ORDER BY seq", (wid,))
    wb = db.query_one("SELECT name, language FROM wordbooks WHERE id=?", (wid,))
    return {"wordbook": wb, "words": words}


@app.get("/api/classes/{cid}")
def class_detail(cid: int, user=Depends(current_user)):
    """선생님용 반 상세: 명단 + 배정 시험지 + 학생별 점수 매트릭스."""
    cls = db.query_one("SELECT * FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    assignments = db.query_all(
        "SELECT a.id, a.exam_id, e.name AS exam_name, e.question_count "
        "FROM assignments a JOIN exams e ON e.id=a.exam_id WHERE a.class_id=? ORDER BY a.assigned_at", (cid,))
    wb_assignments = db.query_all(
        "SELECT wa.id, wa.wordbook_id, w.name AS wordbook_name, w.language, "
        "  (SELECT COUNT(*) FROM words wd WHERE wd.wordbook_id=w.id) AS word_count "
        "FROM wordbook_assignments wa JOIN wordbooks w ON w.id=wa.wordbook_id "
        "WHERE wa.class_id=? ORDER BY wa.assigned_at", (cid,))
    roster = db.query_all(
        "SELECT u.id AS student_id, u.nickname, u.email, rs.sid AS student_sid "
        "FROM class_members m JOIN users u ON u.id=m.student_id "
        "LEFT JOIN roster_students rs ON rs.user_id=u.id AND rs.class_id=m.class_id "
        "WHERE m.class_id=? ORDER BY m.joined_at", (cid,))
    scores = {}
    if assignments and roster:
        exam_ids = [a["exam_id"] for a in assignments]
        sids = [r["student_id"] for r in roster]
        rows = db.query_all(
            f"SELECT user_id, exam_id, MAX(score) best, COUNT(*) attempts FROM attempts "
            f"WHERE exam_id IN ({','.join('?'*len(exam_ids))}) "
            f"AND user_id IN ({','.join('?'*len(sids))}) GROUP BY user_id, exam_id",
            (*exam_ids, *sids))
        for r in rows:
            scores.setdefault(r["user_id"], {})[r["exam_id"]] = {"best": r["best"], "attempts": r["attempts"]}
    return {"class": {"id": cls["id"], "name": cls["name"], "join_code": cls["join_code"]},
            "assignments": assignments, "wordbook_assignments": wb_assignments,
            "roster": roster, "scores": scores}


@app.get("/api/classes/{cid}/students/{sid}")
def class_student_detail(cid: int, sid: int, user=Depends(current_user)):
    """선생님용: 특정 학생의 상세 학습 현황(대시보드 형태)."""
    cls = db.query_one("SELECT id, name FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    member = db.query_one(
        "SELECT u.id, u.nickname, u.email, u.avatar, rs.sid AS student_sid FROM class_members m "
        "JOIN users u ON u.id=m.student_id "
        "LEFT JOIN roster_students rs ON rs.user_id=u.id AND rs.class_id=m.class_id "
        "WHERE m.class_id=? AND m.student_id=?", (cid, sid))
    if not member:
        raise HTTPException(404, "이 반의 학생이 아닙니다.")

    # 이 반에 배정된 시험지 목록
    assignments = db.query_all(
        "SELECT a.exam_id, e.name AS exam_name, e.format, e.question_count "
        "FROM assignments a JOIN exams e ON e.id=a.exam_id WHERE a.class_id=? ORDER BY a.assigned_at", (cid,))

    # 각 과제별 학생의 응시 이력 + 요약 KPI
    per_exam = []
    all_scores = []
    total_attempts = 0
    for a in assignments:
        atts = db.query_all(
            "SELECT id, score, correct, total, time_taken, created_at FROM attempts "
            "WHERE exam_id=? AND user_id=? ORDER BY created_at DESC", (a["exam_id"], sid))
        best = max((x["score"] for x in atts), default=None)
        if best is not None:
            all_scores.append(best)
        total_attempts += len(atts)
        per_exam.append({
            "exam_id": a["exam_id"], "exam_name": a["exam_name"],
            "format_label": ai_quiz.FORMAT_LABELS.get(a["format"], a["format"]),
            "question_count": a["question_count"],
            "best": best, "attempts": atts,
        })

    # 점수 추이 (이 반 과제들에 대한 응시, 오래된→최신)
    exam_ids = [a["exam_id"] for a in assignments]
    trend = []
    if exam_ids:
        trend = db.query_all(
            f"SELECT score, created_at FROM attempts WHERE user_id=? "
            f"AND exam_id IN ({','.join('?'*len(exam_ids))}) ORDER BY created_at DESC LIMIT 20",
            (sid, *exam_ids))
        trend = list(reversed(trend))

    avg = round(sum(all_scores) / len(all_scores)) if all_scores else 0
    best = max(all_scores) if all_scores else 0

    return {
        "class": {"id": cls["id"], "name": cls["name"]},
        "student": {"id": member["id"], "nickname": member["nickname"],
                    "email": member["email"], "avatar": member.get("avatar") or "",
                    "sid": member.get("student_sid") or ""},
        "summary": {"assigned": len(assignments), "completed": len([e for e in per_exam if e["best"] is not None]),
                    "attempts": total_attempts, "avg": avg, "best": best},
        "trend": trend,
        "exams": per_exam,
    }


@app.delete("/api/classes/{cid}")
def delete_class(cid: int, user=Depends(current_user)):
    cls = db.query_one("SELECT id FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    db.execute("DELETE FROM classes WHERE id=?", (cid,), commit=True)
    return {"status": "success"}


class AssignBody(BaseModel):
    exam_id: int


@app.post("/api/classes/{cid}/assign")
def assign_exam(cid: int, body: AssignBody, user=Depends(current_user)):
    cls = db.query_one("SELECT id, name FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    exam = db.query_one("SELECT id, name FROM exams WHERE id=? AND user_id=?", (body.exam_id, user["id"]))
    if not exam:
        raise HTTPException(404, "본인 시험지만 배정할 수 있습니다.")
    existing = db.query_one("SELECT id FROM assignments WHERE class_id=? AND exam_id=?", (cid, body.exam_id))
    if existing:
        return {"status": "success", "assignment_id": existing["id"]}
    aid = db.execute("INSERT INTO assignments (class_id, exam_id, assigned_at) VALUES (?, ?, ?)",
                     (cid, body.exam_id, _now()), commit=True)
    # 반 학생들에게 새 과제 알림 발송
    members = db.query_all("SELECT student_id FROM class_members WHERE class_id=?", (cid,))
    for m in members:
        _notify(m["student_id"], f"새 시험지가 배정됐어요 📝",
                f"'{cls['name']}' 반에 '{exam['name']}' 시험지가 배정되었습니다.", ntype="assign")
    return {"status": "success", "assignment_id": aid}


@app.delete("/api/assignments/{aid}")
def unassign(aid: int, user=Depends(current_user)):
    row = db.query_one(
        "SELECT a.id FROM assignments a JOIN classes c ON c.id=a.class_id "
        "WHERE a.id=? AND c.teacher_id=?", (aid, user["id"]))
    if not row:
        raise HTTPException(404, "과제를 찾을 수 없습니다.")
    db.execute("DELETE FROM assignments WHERE id=?", (aid,), commit=True)
    return {"status": "success"}


class AssignWbBody(BaseModel):
    wordbook_id: int


@app.post("/api/classes/{cid}/assign-wordbook")
def assign_wordbook(cid: int, body: AssignWbBody, user=Depends(current_user)):
    """반에 단어장 배정 → 학생에게 복사본이 생기고 알림."""
    cls = db.query_one("SELECT id, name FROM classes WHERE id=? AND teacher_id=?", (cid, user["id"]))
    if not cls:
        raise HTTPException(404, "반을 찾을 수 없습니다.")
    wb = db.query_one("SELECT id, name FROM wordbooks WHERE id=? AND user_id=?", (body.wordbook_id, user["id"]))
    if not wb:
        raise HTTPException(404, "본인 단어장만 배정할 수 있습니다.")
    existing = db.query_one("SELECT id FROM wordbook_assignments WHERE class_id=? AND wordbook_id=?",
                            (cid, body.wordbook_id))
    if existing:
        return {"status": "success", "assignment_id": existing["id"]}
    aid = db.execute("INSERT INTO wordbook_assignments (class_id, wordbook_id, assigned_at) VALUES (?, ?, ?)",
                     (cid, body.wordbook_id, _now()), commit=True)
    for m in db.query_all("SELECT student_id FROM class_members WHERE class_id=?", (cid,)):
        _notify(m["student_id"], "새 단어장이 배정됐어요 📘",
                f"'{cls['name']}' 반에 '{wb['name']}' 단어장이 배정되었습니다.", ntype="assign")
    return {"status": "success", "assignment_id": aid}


@app.delete("/api/wordbook-assignments/{aid}")
def unassign_wordbook(aid: int, user=Depends(current_user)):
    row = db.query_one(
        "SELECT wa.id FROM wordbook_assignments wa JOIN classes c ON c.id=wa.class_id "
        "WHERE wa.id=? AND c.teacher_id=?", (aid, user["id"]))
    if not row:
        raise HTTPException(404, "배정을 찾을 수 없습니다.")
    db.execute("DELETE FROM wordbook_assignments WHERE id=?", (aid,), commit=True)
    return {"status": "success"}


class JoinBody(BaseModel):
    code: str


@app.post("/api/classes/join")
def join_class(body: JoinBody, user=Depends(current_user)):
    code = body.code.strip().upper()
    cls = db.query_one("SELECT * FROM classes WHERE join_code=?", (code,))
    if not cls:
        raise HTTPException(404, "참여 코드를 찾을 수 없습니다.")
    if cls["teacher_id"] == user["id"]:
        raise HTTPException(400, "본인이 만든 반에는 참여할 수 없습니다.")
    if not db.query_one("SELECT id FROM class_members WHERE class_id=? AND student_id=?", (cls["id"], user["id"])):
        db.execute("INSERT INTO class_members (class_id, student_id, joined_at) VALUES (?, ?, ?)",
                   (cls["id"], user["id"], _now()), commit=True)
    return {"status": "success", "class_id": cls["id"], "name": cls["name"]}


@app.post("/api/classes/{cid}/leave")
def leave_class(cid: int, user=Depends(current_user)):
    db.execute("DELETE FROM class_members WHERE class_id=? AND student_id=?", (cid, user["id"]), commit=True)
    return {"status": "success"}


@app.get("/api/exams/{exam_id}/attempts")
def list_attempts(exam_id: int, user=Depends(current_user)):
    # 소유/배정된 시험지면 접근 허용, 반환은 '본인' 응시 기록만
    if not _exam_visible_to(user["id"], exam_id):
        raise HTTPException(404, "시험지를 찾을 수 없습니다.")
    return {"attempts": db.query_all(
        "SELECT id, score, correct, total, time_taken, created_at "
        "FROM attempts WHERE exam_id=? AND user_id=? ORDER BY created_at DESC", (exam_id, user["id"]),
    )}


@app.get("/api/attempts/{attempt_id}")
def get_attempt(attempt_id: int, user=Depends(current_user)):
    att = db.query_one("SELECT * FROM attempts WHERE id=? AND user_id=?", (attempt_id, user["id"]))
    if not att:
        raise HTTPException(404, "응시 기록을 찾을 수 없습니다.")
    att["results"] = json.loads(att["results"])
    return {"attempt": att}


# ==================================================================
# PDF 다운로드
# ==================================================================
@app.post("/api/report/pdf")
def download_progress_pdf(user=Depends(current_user)):
    """전체 학습 진도 리포트 PDF."""
    uid = user["id"]
    full = auth.get_user_full(uid)
    agg = db.query_one("SELECT COUNT(*) c, COALESCE(ROUND(AVG(score)),0) avg, COALESCE(MAX(score),0) best "
                       "FROM attempts WHERE user_id=?", (uid,))
    week_ago = (datetime.now(timezone.utc) - _timedelta(days=7)).isoformat()
    week = db.query_one("SELECT COUNT(*) c FROM attempts WHERE user_id=? AND created_at>=?", (uid, week_ago))["c"]
    exams = db.query_all(
        "SELECT e.name, MAX(a.score) best, ROUND(AVG(a.score)) avg, COUNT(a.id) attempts "
        "FROM exams e JOIN attempts a ON a.exam_id=e.id AND a.user_id=? WHERE e.user_id=? "
        "GROUP BY e.id ORDER BY MAX(a.created_at) DESC LIMIT 12", (uid, uid))
    an = get_analysis(user=user)
    data = {
        "name": full.get("nickname") or full["email"].split("@")[0],
        "avg": agg["avg"], "best": agg["best"], "attempts": agg["c"], "week": week,
        "mastery": an.get("distribution", {}), "exams": exams,
        "weak_words": an.get("weak_words", []),
    }
    out = os.path.join(UPLOAD_DIR, f"progress_{uid}.pdf")
    pdf_export.create_progress_pdf(data, out)
    return FileResponse(out, filename="학습리포트.pdf", media_type="application/pdf")


class PdfOptions(BaseModel):
    title: str | None = None
    font: str = "round"        # round(둥근고딕) | serif(명조) | sans(기본고딕)
    font_size: int = 10        # 문항 글자 크기(pt)
    columns: int = 1           # 1 | 2 단
    spacing: int = 8           # 문항 간격
    answer_key: bool = True     # (사용 안함) 정답지는 별도 파일로
    header_fields: bool = True  # 이름/날짜/점수 칸
    show_school: bool = True    # 상단 브랜드 문구
    kind: str = "quiz"          # quiz(문제지) | answers(정답·해설지, 별도 PDF)


@app.post("/api/exams/{exam_id}/pdf")
def download_exam_pdf(exam_id: int, user=Depends(current_user), opts: PdfOptions | None = None):
    exam = db.query_one("SELECT * FROM exams WHERE id=? AND user_id=?", (exam_id, user["id"]))
    if not exam:
        raise HTTPException(404, "시험지를 찾을 수 없습니다.")
    o = (opts or PdfOptions()).dict()
    kind = o.get("kind", "quiz")
    req_title = o.get("title")
    full = auth.get_user_full(user["id"])
    lim = limits_for(full.get("tier"))
    # 무료 등급: 옵션은 기본값으로 되돌리되(제목·kind 는 유지), 워터마크 강제
    if not lim.get("pdf_custom"):
        o = PdfOptions().dict()
        o["title"] = req_title
    o["kind"] = kind
    o["watermark"] = lim.get("pdf_watermark", False)
    o["title"] = (o.get("title") or exam["name"]).strip() or exam["name"]
    suffix = "_answers" if kind == "answers" else ""
    out = os.path.join(UPLOAD_DIR, f"quiz_{exam_id}{suffix}.pdf")
    pdf_export.create_quiz_pdf(json.loads(exam["questions"]), out, options=o)
    fname = f"{o['title']}{'_정답' if kind == 'answers' else ''}.pdf"
    return FileResponse(out, filename=fname, media_type="application/pdf")


@app.post("/api/attempts/{attempt_id}/pdf")
def download_report_pdf(attempt_id: int, user=Depends(current_user)):
    att = db.query_one("SELECT * FROM attempts WHERE id=? AND user_id=?", (attempt_id, user["id"]))
    if not att:
        raise HTTPException(404, "응시 기록을 찾을 수 없습니다.")
    data = {"score": att["score"], "correct": att["correct"], "total": att["total"],
            "results": json.loads(att["results"])}
    out = os.path.join(UPLOAD_DIR, f"report_{attempt_id}.pdf")
    pdf_export.create_report_pdf(data, out)
    return FileResponse(out, filename=f"report_{attempt_id}.pdf", media_type="application/pdf")


# ==================================================================
# 정적 파일 (프론트엔드)
# ==================================================================
app.mount("/", StaticFiles(directory="static", html=True), name="static")
